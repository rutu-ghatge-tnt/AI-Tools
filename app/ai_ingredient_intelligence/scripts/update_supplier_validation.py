"""
Script to update supplier validation flags based on Excel file highlighting.
Marks highlighted suppliers as valid (isValid: true) and others as invalid (isValid: false).
"""
import pandas as pd
from pathlib import Path
import asyncio
from motor.motor_asyncio import AsyncIOMotorClient
from bson import ObjectId
from app.config import MONGO_URI, DB_NAME

# Excel file path
SCRIPT_DIR = Path(__file__).parent
ROOT_DIR = SCRIPT_DIR.parent.parent.parent
EXCEL_FILE = ROOT_DIR / "skin_bb.ingre_suppliers_updated (2).xlsx"

if not EXCEL_FILE.exists():
    EXCEL_FILE = Path("skin_bb.ingre_suppliers_updated (2).xlsx")

try:
    import openpyxl
    from openpyxl.styles import PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not available - will use pandas only")


def get_highlighted_suppliers_from_excel():
    """
    Read Excel file and identify suppliers with yellow highlighted cells.
    Returns a set of highlighted supplier names.
    """
    highlighted_suppliers = set()
    all_suppliers = set()
    
    if not OPENPYXL_AVAILABLE:
        # Fallback: read with pandas only
        try:
            df = pd.read_excel(EXCEL_FILE)
            print(f"Excel file has {len(df)} rows")
            print(f"Columns: {list(df.columns)}")
            # Find supplier column
            supplier_col = None
            for col in df.columns:
                col_lower = str(col).lower()
                if 'supplier' in col_lower:
                    supplier_col = col
                    break
            if supplier_col:
                all_suppliers = set(df[supplier_col].dropna().astype(str).str.strip().unique())
                print(f"Found {len(all_suppliers)} suppliers in Excel (highlighting detection unavailable)")
                return set(), all_suppliers
            return set(), set()
        except Exception as e:
            print(f"Error reading Excel: {e}")
            return set(), set()
    
    try:
        # Load workbook with openpyxl to access cell formatting
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        sheet = wb.active
        
        # Check all cells for yellow fill
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    cell_value = str(cell.value).strip()
                    if cell_value:
                        # Check if cell has yellow fill
                        if cell.fill and cell.fill.fill_type == "solid":
                            fill_color = cell.fill.start_color
                            if fill_color:
                                rgb = fill_color.rgb
                                if rgb:
                                    # Yellow is typically FFFF00 or FFFFFF00
                                    if "FFFF00" in str(rgb).upper() or "FFFFFF00" in str(rgb).upper():
                                        highlighted_suppliers.add(cell_value)
                                # Also check indexed colors
                                elif hasattr(fill_color, 'index') and fill_color.index in [6, 13]:
                                    highlighted_suppliers.add(cell_value)
        
        # Also read with pandas to get all supplier names
        df = pd.read_excel(EXCEL_FILE)
        
        # Find supplier column
        supplier_col = None
        for col in df.columns:
            col_lower = str(col).lower()
            if 'supplier' in col_lower:
                supplier_col = col
                break
        
        if supplier_col:
            all_suppliers = set(df[supplier_col].dropna().astype(str).str.strip().unique())
            
            if highlighted_suppliers:
                print(f"Found {len(highlighted_suppliers)} highlighted suppliers from cell formatting")
            else:
                print("Could not detect highlighted cells automatically")
                print(f"   Found {len(all_suppliers)} total suppliers in Excel")
        
        return highlighted_suppliers, all_suppliers
        
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        # Fallback: try reading with pandas only
        try:
            df = pd.read_excel(EXCEL_FILE)
            print(f"Excel file has {len(df)} rows")
            print(f"Columns: {list(df.columns)}")
            return set(), set()
        except Exception as e2:
            print(f"Error reading Excel with pandas: {e2}")
            return set(), set()


async def update_supplier_validation():
    """
    Update supplier validation flags based on Excel highlighting.
    """
    client = AsyncIOMotorClient(MONGO_URI)
    db = client[DB_NAME]
    
    suppliers_col = db["ingre_suppliers"]
    
    # Get highlighted suppliers from Excel
    highlighted_suppliers, all_excel_suppliers = get_highlighted_suppliers_from_excel()
    
    if not highlighted_suppliers:
        print("\n" + "="*60)
        print("ERROR: Could not detect highlighted suppliers from Excel")
        print("="*60)
        print("Please ensure the Excel file has yellow highlighted supplier names.")
        print("="*60 + "\n")
        return
    
    # Get all suppliers from database
    all_suppliers_cursor = suppliers_col.find({}, {"supplierName": 1, "_id": 1, "isValid": 1})
    all_suppliers = await all_suppliers_cursor.to_list(length=None)
    
    # Create mapping for case-insensitive matching
    supplier_name_to_doc = {}
    for s in all_suppliers:
        name = s.get("supplierName", "").strip()
        if name:
            supplier_name_to_doc[name.lower()] = s
    
    # Match highlighted suppliers with database suppliers
    highlighted_supplier_ids = set()
    matched_names = []
    unmatched_highlighted = []
    
    for supplier_name in highlighted_suppliers:
        supplier_name_clean = supplier_name.strip()
        # Try exact match first
        if supplier_name_clean in supplier_name_to_doc:
            doc = supplier_name_to_doc[supplier_name_clean]
            highlighted_supplier_ids.add(str(doc["_id"]))
            matched_names.append(supplier_name_clean)
        else:
            # Try case-insensitive match
            found = False
            for db_name_lower, doc in supplier_name_to_doc.items():
                if db_name_lower == supplier_name_clean.lower():
                    highlighted_supplier_ids.add(str(doc["_id"]))
                    matched_names.append(doc.get("supplierName", ""))
                    found = True
                    break
            if not found:
                unmatched_highlighted.append(supplier_name_clean)
    
    print(f"\nMatched {len(matched_names)} highlighted suppliers from Excel to database")
    if unmatched_highlighted:
        print(f"Warning: {len(unmatched_highlighted)} highlighted suppliers from Excel not found in database:")
        for name in unmatched_highlighted[:10]:  # Show first 10
            try:
                print(f"  - {name}")
            except UnicodeEncodeError:
                print(f"  - {name.encode('ascii', 'ignore').decode('ascii')}")
        if len(unmatched_highlighted) > 10:
            print(f"  ... and {len(unmatched_highlighted) - 10} more")
    
    # Update all suppliers
    # First, mark all as invalid (default)
    result_invalid = await suppliers_col.update_many(
        {},
        {"$set": {"isValid": False}}
    )
    print(f"\nMarked {result_invalid.modified_count} suppliers as invalid (isValid: false)")
    
    # Then, mark highlighted suppliers as valid
    if highlighted_supplier_ids:
        object_ids = [ObjectId(id) for id in highlighted_supplier_ids]
        result_valid = await suppliers_col.update_many(
            {"_id": {"$in": object_ids}},
            {"$set": {"isValid": True}}
        )
        print(f"Marked {result_valid.modified_count} suppliers as valid (isValid: true)")
    
    # Verify counts
    valid_count = await suppliers_col.count_documents({"isValid": True})
    invalid_count = await suppliers_col.count_documents({"isValid": False})
    null_count = await suppliers_col.count_documents({"isValid": {"$exists": False}})
    
    print("\n" + "="*60)
    print("UPDATE SUMMARY")
    print("="*60)
    print(f"Valid Suppliers (isValid: true): {valid_count}")
    print(f"Invalid Suppliers (isValid: false): {invalid_count}")
    if null_count > 0:
        print(f"Suppliers without isValid field: {null_count}")
    print("="*60 + "\n")
    
    # Count ingredients
    branded_ingredients_col = db["ingre_branded_ingredients"]
    
    # Count ingredients with valid suppliers
    valid_ingredients = await branded_ingredients_col.count_documents({
        "supplier_id": {"$in": [ObjectId(id) for id in highlighted_supplier_ids]}
    })
    
    # Count total ingredients
    total_ingredients = await branded_ingredients_col.count_documents({})
    
    print("INGREDIENT IMPACT:")
    print(f"Total Branded Ingredients: {total_ingredients}")
    print(f"Ingredients with Valid Suppliers: {valid_ingredients}")
    print(f"Ingredients that will be HIDDEN: {total_ingredients - valid_ingredients}")
    print("="*60 + "\n")
    
    client.close()


if __name__ == "__main__":
    asyncio.run(update_supplier_validation())

