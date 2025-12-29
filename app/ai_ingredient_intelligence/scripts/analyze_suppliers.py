"""
Script to analyze suppliers from Excel file and count highlighted suppliers
and their linked products/ingredients.
"""
import pandas as pd
from pathlib import Path
import asyncio
from motor.motor_asyncio import AsyncIOMotorClient
from bson import ObjectId
from app.config import MONGO_URI, DB_NAME

try:
    import openpyxl
    from openpyxl.styles import PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️  openpyxl not available - will use pandas only")

# Excel file path - check root directory
SCRIPT_DIR = Path(__file__).parent
ROOT_DIR = SCRIPT_DIR.parent.parent.parent  # Go up from scripts -> logic -> ai_ingredient_intelligence -> app -> root
EXCEL_FILE = ROOT_DIR / "skin_bb.ingre_suppliers_updated (2).xlsx"

# If not found in root, try current directory
if not EXCEL_FILE.exists():
    EXCEL_FILE = Path("skin_bb.ingre_suppliers_updated (2).xlsx")

def get_highlighted_suppliers_from_excel():
    """
    Read Excel file and identify suppliers with yellow highlighted cells.
    Returns a set of highlighted supplier names.
    """
    highlighted_suppliers = set()
    
    if not OPENPYXL_AVAILABLE:
        # Fallback: read with pandas only
        try:
            df = pd.read_excel(EXCEL_FILE)
            print(f"📊 Excel file has {len(df)} rows")
            print(f"📊 Columns: {list(df.columns)}")
            # Find supplier column
            supplier_col = None
            for col in df.columns:
                col_lower = str(col).lower()
                if 'supplier' in col_lower:
                    supplier_col = col
                    break
            if supplier_col:
                all_suppliers = set(df[supplier_col].dropna().astype(str).str.strip().unique())
                return set(), all_suppliers
            return set(), set()
        except Exception as e:
            print(f"Error reading Excel: {e}")
            return set(), set()
    
    try:
        # Load workbook with openpyxl to access cell formatting
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        
        # Try to find the sheet with supplier data
        # Usually it's the first sheet or named 'Sheet1'
        sheet = wb.active
        
        # Yellow fill pattern (common highlight color)
        yellow_fills = [
            PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid"),  # Yellow
            PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),  # Yellow (hex)
            PatternFill(start_color="FFFF00FF", end_color="FFFF00FF", fill_type="solid"),  # Yellow (ARGB)
        ]
        
        # Check all cells for yellow fill
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    cell_value = str(cell.value).strip()
                    if cell_value:
                        # Check if cell has yellow fill
                        if cell.fill and cell.fill.fill_type == "solid":
                            fill_color = cell.fill.start_color
                            # Check if it's yellow (RGB: 255, 255, 0 or similar)
                            if fill_color:
                                rgb = fill_color.rgb
                                if rgb:
                                    # Yellow is typically FFFF00 or FFFFFF00
                                    if "FFFF00" in str(rgb).upper() or "FFFFFF00" in str(rgb).upper():
                                        highlighted_suppliers.add(cell_value)
                                # Also check indexed colors
                                elif hasattr(fill_color, 'index') and fill_color.index in [6, 13]:  # Common yellow indices
                                    highlighted_suppliers.add(cell_value)
        
        # Also try reading with pandas to get all supplier names
        df = pd.read_excel(EXCEL_FILE)
        
        # Find supplier column (common names: supplier, supplierName, supplier_name, etc.)
        supplier_col = None
        for col in df.columns:
            col_lower = str(col).lower()
            if 'supplier' in col_lower:
                supplier_col = col
                break
        
        if supplier_col:
            # Get all unique suppliers from the column
            all_suppliers = set(df[supplier_col].dropna().astype(str).str.strip().unique())
            
            # If we found highlighted suppliers, use them
            # Otherwise, we'll need to manually check which ones are highlighted
            if highlighted_suppliers:
                print(f"Found {len(highlighted_suppliers)} highlighted suppliers from cell formatting")
            else:
                print("Could not detect highlighted cells automatically")
                print(f"   Found {len(all_suppliers)} total suppliers in Excel")
                print("   Please check the Excel file manually or provide a list of highlighted supplier names")
        
        return highlighted_suppliers, all_suppliers if supplier_col else set()
        
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        # Fallback: try reading with pandas only
        try:
            df = pd.read_excel(EXCEL_FILE)
            print(f"Excel file has {len(df)} rows")
            print(f"Columns: {list(df.columns)}")
            return set(), set()
        except Exception as e2:
            print(f"Error reading Excel with pandas: {e2}")
            return set(), set()


async def count_suppliers_and_products():
    """
    Count suppliers in database and products linked to highlighted suppliers.
    """
    client = AsyncIOMotorClient(MONGO_URI)
    db = client[DB_NAME]
    
    suppliers_col = db["ingre_suppliers"]
    branded_ingredients_col = db["ingre_branded_ingredients"]
    
    # Get all suppliers from database
    all_suppliers_cursor = suppliers_col.find({}, {"supplierName": 1, "_id": 1})
    all_suppliers = await all_suppliers_cursor.to_list(length=None)
    
    # Create mapping: supplier_id -> supplier_name
    supplier_id_to_name = {str(s["_id"]): s.get("supplierName", "") for s in all_suppliers}
    supplier_name_to_id = {s.get("supplierName", ""): str(s["_id"]) for s in all_suppliers}
    
    total_suppliers = len(all_suppliers)
    
    # Get highlighted suppliers from Excel
    highlighted_suppliers, all_excel_suppliers = get_highlighted_suppliers_from_excel()
    
    # If we couldn't detect highlights automatically, show all suppliers
    if not highlighted_suppliers and all_excel_suppliers:
        print("\n" + "="*60)
        print("COULD NOT AUTO-DETECT HIGHLIGHTED SUPPLIERS")
        print("="*60)
        print(f"Found {len(all_excel_suppliers)} suppliers in Excel file")
        print("\nPlease provide the list of highlighted supplier names,")
        print("or we can mark all suppliers as 'valid' by default.")
        print("="*60 + "\n")
    
    # Match highlighted suppliers with database suppliers
    highlighted_supplier_ids = set()
    for supplier_name in highlighted_suppliers:
        # Try exact match
        if supplier_name in supplier_name_to_id:
            highlighted_supplier_ids.add(supplier_name_to_id[supplier_name])
        else:
            # Try case-insensitive match
            for db_name, db_id in supplier_name_to_id.items():
                if db_name.lower() == supplier_name.lower():
                    highlighted_supplier_ids.add(db_id)
                    break
    
    highlighted_count = len(highlighted_supplier_ids)
    non_highlighted_count = total_suppliers - highlighted_count
    
    # Count products/ingredients linked to highlighted suppliers
    highlighted_ingredients = await branded_ingredients_col.count_documents({
        "supplier_id": {"$in": [ObjectId(id) for id in highlighted_supplier_ids]}
    })
    
    # Count total branded ingredients
    total_ingredients = await branded_ingredients_col.count_documents({})
    
    # Count ingredients with any supplier
    ingredients_with_supplier = await branded_ingredients_col.count_documents({
        "supplier_id": {"$ne": None}
    })
    
    # Count ingredients with non-highlighted suppliers
    non_highlighted_ingredients = ingredients_with_supplier - highlighted_ingredients
    
    print("\n" + "="*60)
    print("SUPPLIER ANALYSIS REPORT")
    print("="*60)
    print(f"Total Suppliers in Database: {total_suppliers}")
    print(f"Highlighted Suppliers (from Excel): {highlighted_count}")
    print(f"Non-Highlighted Suppliers: {non_highlighted_count}")
    print("\n" + "-"*60)
    print(f"Total Branded Ingredients: {total_ingredients}")
    print(f"Ingredients with Any Supplier: {ingredients_with_supplier}")
    print(f"Ingredients Linked to Highlighted Suppliers: {highlighted_ingredients}")
    print(f"Ingredients Linked to Non-Highlighted Suppliers: {non_highlighted_ingredients}")
    print(f"Ingredients with No Supplier: {total_ingredients - ingredients_with_supplier}")
    print("="*60 + "\n")
    
    return {
        "total_suppliers": total_suppliers,
        "highlighted_suppliers": highlighted_count,
        "non_highlighted_suppliers": non_highlighted_count,
        "highlighted_supplier_ids": highlighted_supplier_ids,
        "highlighted_supplier_names": highlighted_suppliers,
        "total_ingredients": total_ingredients,
        "ingredients_with_supplier": ingredients_with_supplier,
        "highlighted_ingredients": highlighted_ingredients,
        "non_highlighted_ingredients": non_highlighted_ingredients
    }


if __name__ == "__main__":
    asyncio.run(count_suppliers_and_products())

